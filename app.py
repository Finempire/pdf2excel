#!/usr/bin/env python3
"""
Universal Bank Statement PDF -> Excel converter.

Designed for common Indian bank statements (including BoB-style Internet statements).

Extraction strategy (in order):
1) Camelot (if installed) - tries to extract real tables (best when it works).
2) Tabula (if installed)  - table extraction using Java.
3) pdfplumber line parsing - robust for many "text PDFs" even when tables don't extract cleanly.
4) Optional OCR fallback   - only if you enable --ocr (requires external system deps).

Output Excel:
- Summary (metadata + extraction info)
- Transactions (normalized transactions)
- Raw (optional, helpful for debugging)

USAGE
-----
pip install pdfplumber pandas openpyxl
# optional (tables):
pip install "camelot-py[cv]"    # requires ghostscript + opencv
pip install tabula-py           # requires Java
# optional (ocr):
pip install pytesseract pdf2image  # requires tesseract + poppler

python pdf_to_excel_bank_statement_universal.py input.pdf output.xlsx
python pdf_to_excel_bank_statement_universal.py input.pdf output.xlsx --method auto
python pdf_to_excel_bank_statement_universal.py input.pdf output.xlsx --method lines
python pdf_to_excel_bank_statement_universal.py input.pdf output.xlsx --method camelot
python pdf_to_excel_bank_statement_universal.py input.pdf output.xlsx --ocr

NOTES
-----
- This script assumes a "running balance" exists to infer DR/CR when the PDF text
  doesn't clearly preserve debit/credit columns.
- OCR is off by default because it's slower and less accurate than text extraction.
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import os
import re
import sys
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd

# --- core libs for the "always-on" method ---
import pdfplumber

# --- optional libs (imported lazily inside functions) ---
# camelot, tabula, pytesseract, pdf2image


# -----------------------------
# Regex and parsing helpers
# -----------------------------
# Accepts "14/11/2024", "14-11-2024", "14/Nov/2024" (with or without stray spaces/newlines)
DATE_RE = re.compile(r"^(\d{2})\s*[/-]\s*(\d{2}|[A-Za-z]{3})\s*[/-]\s*(\d{4})\b", re.IGNORECASE)
# Balance token in BoB often looks like "83,47,632.52Cr" or "83,47,632.52 Cr"
BAL_TOKEN_RE = re.compile(r"(\d[\d,]*\.\d{2})\s*(Cr|Dr)\b", re.IGNORECASE)
# Generic amount token (strict: must have decimal to avoid matching reference IDs)
AMT_TOKEN_RE = re.compile(r"\b(\d[\d,]*\.\d{2})\b")


def parse_float_indian(num_str: str) -> float:
    """Convert '1,23,456.78' -> 123456.78"""
    return float(num_str.replace(",", ""))


def safe_strip(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def normalize_whitespace(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def match_date_token(s: str) -> Optional[re.Match]:
    """Return DATE_RE match after stripping whitespace/newlines inside the token."""
    return DATE_RE.match(re.sub(r"\s+", "", safe_strip(s)))


def parse_date(d: str) -> Optional[dt.date]:
    d_clean = re.sub(r"\s+", "", d.strip())
    m = DATE_RE.match(d_clean)
    if not m:
        return None
    day = m.group(1)
    month_part = m.group(2)
    year = m.group(3)
    try:
        if month_part.isalpha():
            return dt.datetime.strptime(f"{day}{month_part}{year}", "%d%b%Y").date()
        return dt.datetime.strptime(f"{day}{month_part}{year}", "%d%m%Y").date()
    except Exception:
        return None


# -----------------------------
# Data model
# -----------------------------
@dataclass
class Txn:
    date: dt.date
    narration: str
    withdrawal_dr: Optional[float]
    deposit_cr: Optional[float]
    balance: Optional[float]
    balance_type: Optional[str]  # 'Cr'/'Dr'/None
    inferred_type: Optional[str]  # 'DR'/'CR'/None
    raw_amounts: List[float]
    source: str  # 'lines'/'camelot'/'tabula'/'ocr'
    raw_text: str = ""


# -----------------------------
# Metadata extraction
# -----------------------------
def extract_metadata(full_text: str) -> Dict[str, Optional[str]]:
    """
    Pull basic metadata if present (BoB style patterns).
    Works as "best effort" for other banks too.
    """
    meta: Dict[str, Optional[str]] = {}

    def _m(pattern: str) -> Optional[re.Match]:
        return re.search(pattern, full_text, flags=re.IGNORECASE)

    m = _m(r"Customer\s*Id:\s*([A-Z0-9]+)")
    meta["Customer ID"] = m.group(1).strip() if m else None

    m = _m(r"Branch\s*Name:\s*(.+)")
    meta["Branch"] = m.group(1).strip() if m else None

    m = _m(r"IFSC\s*Code:\s*([A-Z0-9]+)")
    meta["IFSC"] = m.group(1).strip() if m else None

    m = _m(r"Account\s*No:\s*([0-9X]+)")
    meta["Account No"] = m.group(1).strip() if m else None

    m = _m(r"Statement\s*Period\s*from\s*(\d{2}[/-]\d{2}[/-]\d{4})\s*to\s*(\d{2}[/-]\d{2}[/-]\d{4})")
    meta["Statement Period"] = f"{m.group(1)} to {m.group(2)}" if m else None

    # Some statements mention "Your Account Statement as on ..."
    m = _m(r"Statement\s*as\s*on\s*(\d{2}[/-]\d{2}[/-]\d{4})")
    meta["Statement As On"] = m.group(1) if m else None

    return meta


# -----------------------------
# Method 1/2: Table extraction
# -----------------------------
def try_camelot_tables(pdf_path: str) -> List[pd.DataFrame]:
    """
    Return list of DataFrames extracted by Camelot.
    Requires: camelot-py and (often) ghostscript.
    """
    try:
        import camelot  # type: ignore
    except Exception:
        return []

    dfs: List[pd.DataFrame] = []
    # Try both flavors: 'lattice' works when ruled lines exist, 'stream' for whitespace tables.
    for flavor in ("lattice", "stream"):
        try:
            tables = camelot.read_pdf(pdf_path, pages="all", flavor=flavor)
            for t in tables:
                df = t.df
                if df is not None and not df.empty:
                    dfs.append(df)
        except Exception:
            continue
    return dfs


def try_tabula_tables(pdf_path: str) -> List[pd.DataFrame]:
    """
    Return list of DataFrames extracted by tabula-py.
    Requires: tabula-py + Java runtime.
    """
    try:
        import tabula  # type: ignore
    except Exception:
        return []

    dfs: List[pd.DataFrame] = []
    try:
        out = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, stream=True)
        if isinstance(out, list):
            dfs.extend([df for df in out if df is not None and not df.empty])
    except Exception:
        pass

    try:
        out = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, lattice=True)
        if isinstance(out, list):
            dfs.extend([df for df in out if df is not None and not df.empty])
    except Exception:
        pass

    return dfs


def looks_like_txn_table(df: pd.DataFrame) -> bool:
    """
    Heuristic: Does the DF contain many rows with a date-like first column?
    """
    if df is None or df.empty:
        return False
    if len(df) < 5:
        return False

    best_ratio = 0.0
    # Check a few leading columns to spot a date-like column
    for col in df.columns[:4]:
        ser = df[col].astype(str)
        values = list(ser.head(30))
        if not values:
            continue
        hits = sum(1 for v in values if match_date_token(str(v)))
        ratio = hits / len(values)
        best_ratio = max(best_ratio, ratio)
    return best_ratio >= 0.20


def clean_table_df(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    df2 = df2.apply(lambda col: col.map(lambda x: normalize_whitespace(str(x).replace("\n", " ")) if x is not None else ""))
    first_row = df2.iloc[0].astype(str).str.lower().tolist()
    header_words = {"date", "narration", "particular", "description", "withdrawal", "debit", "deposit", "credit", "balance", "chq", "cheque"}
    if any(any(w in cell for w in header_words) for cell in first_row):
        df2 = df2.iloc[1:].reset_index(drop=True)
    return df2


def normalize_table_to_txns(df: pd.DataFrame, source: str) -> List[Txn]:
    df = clean_table_df(df)

    cols = [safe_strip(c) for c in df.columns]
    sample_row = df.iloc[0].astype(str).str.lower().tolist() if len(df) else []
    merged_headers = [safe_strip(c).lower() for c in cols]

    best_date_col = None
    best_date_hits = 0
    for c in df.columns:
        hits = 0
        for v in df[c].astype(str).head(50):
            if match_date_token(str(v)):
                hits += 1
        if hits > best_date_hits:
            best_date_hits = hits
            best_date_col = c

    if best_date_col is None or best_date_hits < 2:
        return []

    other_cols = [c for c in df.columns if c != best_date_col]
    if not other_cols:
        return []

    best_bal_col = None
    best_bal_hits = 0
    for c in other_cols:
        hits = 0
        for v in df[c].astype(str).head(80):
            s_compact = re.sub(r"\s+", "", str(v))
            if BAL_TOKEN_RE.search(s_compact) or AMT_TOKEN_RE.search(s_compact):
                hits += 1
        if hits > best_bal_hits:
            best_bal_hits = hits
            best_bal_col = c

    def avg_len(c: Any) -> float:
        ser = df[c].astype(str).head(80)
        return float(ser.map(lambda x: len(str(x))).mean())

    narr_candidates = [c for c in other_cols if c != best_bal_col]
    narration_col = max(narr_candidates, key=avg_len) if narr_candidates else other_cols[0]

    debit_col = None
    credit_col = None

    header_map = {c: safe_strip(c).lower() for c in df.columns}
    for c in df.columns:
        name = header_map[c]
        if any(k in name for k in ["withdrawal", "debit", "dr"]):
            debit_col = c
        if any(k in name for k in ["deposit", "credit", "cr"]):
            credit_col = c
        if any(k in name for k in ["balance"]):
            best_bal_col = c

    if debit_col is None or credit_col is None:
        if len(df.columns) >= 5:
            cols_list = list(df.columns)
            bal_guess = cols_list[-1]
            d_guess = cols_list[-3] if len(cols_list) >= 3 else None
            c_guess = cols_list[-2] if len(cols_list) >= 2 else None

            def numeric_rate(col: Any) -> float:
                ser = df[col].astype(str).head(80)
                hits = 0
                total = 0
                for v in ser:
                    total += 1
                    if AMT_TOKEN_RE.search(re.sub(r"\s+", "", str(v))):
                        hits += 1
                return hits / total if total else 0.0

            if best_bal_col is None and numeric_rate(bal_guess) >= 0.20:
                best_bal_col = bal_guess
            if d_guess is not None and numeric_rate(d_guess) >= 0.05 and d_guess != best_bal_col:
                debit_col = d_guess
            if c_guess is not None and numeric_rate(c_guess) >= 0.05 and c_guess != best_bal_col:
                credit_col = c_guess

    txns: List[Txn] = []
    for _, r in df.iterrows():
        date_raw = safe_strip(r.get(best_date_col))
        narration = normalize_whitespace(safe_strip(r.get(narration_col)))

        m_date = match_date_token(date_raw)
        if not m_date:
            # Treat as continuation of previous txn narration when date is absent.
            if txns and narration:
                txns[-1].narration = normalize_whitespace(f"{txns[-1].narration} {narration}")
            continue

        d = parse_date(m_date.group(0))  # type: ignore
        if d is None:
            continue

        bal_raw = safe_strip(r.get(best_bal_col)) if best_bal_col is not None else ""
        bal_raw_compact = bal_raw.replace(" ", "")

        bal_val: Optional[float] = None
        bal_type: Optional[str] = None
        m_bal = BAL_TOKEN_RE.search(bal_raw_compact)
        if m_bal:
            bal_val = parse_float_indian(m_bal.group(1))
            bal_type = m_bal.group(2).title()
        else:
            m_num = AMT_TOKEN_RE.search(bal_raw_compact)
            if m_num:
                bal_val = parse_float_indian(m_num.group(1))
                bal_type = None

        def parse_amt_cell(x: Any) -> Optional[float]:
            s = safe_strip(x)
            s_compact = re.sub(r"\s+", "", s)
            m = AMT_TOKEN_RE.search(s_compact)
            return parse_float_indian(m.group(1)) if m else None

        dr = parse_amt_cell(r.get(debit_col)) if debit_col is not None else None
        cr = parse_amt_cell(r.get(credit_col)) if credit_col is not None else None

        raw_amounts: List[float] = []
        for c in df.columns:
            if c == best_bal_col:
                continue
            m = AMT_TOKEN_RE.search(re.sub(r"\s+", "", safe_strip(r.get(c))))
            if m:
                raw_amounts.append(parse_float_indian(m.group(1)))

        txns.append(
            Txn(
                date=d,
                narration=narration,
                withdrawal_dr=dr,
                deposit_cr=cr,
                balance=bal_val,
                balance_type=bal_type,
                inferred_type=None,
                raw_amounts=raw_amounts,
                source=source,
                raw_text="",
            )
        )

    return txns


# -----------------------------
# Method 3: Line-based parsing
# -----------------------------
def extract_lines_pdfplumber(pdf_path: str) -> List[str]:
    lines: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for ln in text.splitlines():
                ln = ln.strip()
                if ln:
                    lines.append(ln)
    return lines


def build_date_blocks(lines: Sequence[str]) -> List[List[str]]:
    blocks: List[List[str]] = []
    current: List[str] = []

    for ln in lines:
        if match_date_token(ln):
            if current:
                blocks.append(current)
            current = [ln]
        else:
            if current:
                current.append(ln)

    if current:
        blocks.append(current)
    return blocks


def parse_block_to_txn(block: Sequence[str], source: str = "lines") -> Optional[Txn]:
    if not block:
        return None

    date_token = block[0]
    m_date = match_date_token(date_token)
    if not m_date:
        return None
    d = parse_date(m_date.group(0))
    if d is None:
        return None
    rest = normalize_whitespace(" ".join(block[1:]))

    bal_val: Optional[float] = None
    bal_type: Optional[str] = None
    bal_span: Optional[Tuple[int, int]] = None

    bal_matches = list(BAL_TOKEN_RE.finditer(rest))
    if bal_matches:
        m = bal_matches[-1]
        bal_val = parse_float_indian(m.group(1))
        bal_type = m.group(2).title()
        bal_span = (m.start(), m.end())
    else:
        amt_matches = list(AMT_TOKEN_RE.finditer(rest))
        if amt_matches:
            m = amt_matches[-1]
            bal_val = parse_float_indian(m.group(1))
            bal_type = None
            bal_span = (m.start(), m.end())

    if bal_span is None:
        return None

    rest_wo_bal = normalize_whitespace((rest[:bal_span[0]] + " " + rest[bal_span[1]:]).strip())

    amt_matches = list(AMT_TOKEN_RE.finditer(rest_wo_bal))
    raw_amounts = [parse_float_indian(m.group(1)) for m in amt_matches]

    narration = rest_wo_bal
    if amt_matches:
        m_last = amt_matches[-1]
        narration = normalize_whitespace((rest_wo_bal[:m_last.start()] + " " + rest_wo_bal[m_last.end():]).strip())

    return Txn(
        date=d,
        narration=narration,
        withdrawal_dr=None,
        deposit_cr=None,
        balance=bal_val,
        balance_type=bal_type,
        inferred_type=None,
        raw_amounts=raw_amounts,
        source=source,
        raw_text=joined,
    )


def parse_lines_method(pdf_path: str) -> Tuple[List[Txn], Dict[str, Any], List[str]]:
    lines = extract_lines_pdfplumber(pdf_path)
    blocks = build_date_blocks(lines)

    txns: List[Txn] = []
    for b in blocks:
        t = parse_block_to_txn(b, source="lines")
        if t:
            txns.append(t)

    txns = infer_dr_cr_from_balance(txns)

    diag = {
        "method": "lines",
        "lines_extracted": len(lines),
        "date_blocks": len(blocks),
        "txns_parsed": len(txns),
    }
    return txns, diag, lines


# -----------------------------
# Method 4: OCR (optional)
# -----------------------------
def ocr_pdf_to_text_lines(pdf_path: str, dpi: int = 300) -> List[str]:
    try:
        from pdf2image import convert_from_path  # type: ignore
        import pytesseract  # type: ignore
    except Exception as e:
        raise RuntimeError("OCR requires pdf2image + pytesseract installed") from e

    pages = convert_from_path(pdf_path, dpi=dpi)
    lines: List[str] = []
    for img in pages:
        text = pytesseract.image_to_string(img)
        for ln in text.splitlines():
            ln = ln.strip()
            if ln:
                lines.append(ln)
    return lines


def parse_ocr_method(pdf_path: str) -> Tuple[List[Txn], Dict[str, Any], List[str]]:
    lines = ocr_pdf_to_text_lines(pdf_path)
    blocks = build_date_blocks(lines)

    txns: List[Txn] = []
    for b in blocks:
        t = parse_block_to_txn(b, source="ocr")
        if t:
            txns.append(t)

    txns = infer_dr_cr_from_balance(txns)

    diag = {
        "method": "ocr",
        "lines_extracted": len(lines),
        "date_blocks": len(blocks),
        "txns_parsed": len(txns),
    }
    return txns, diag, lines


# -----------------------------
# Balance delta inference
# -----------------------------
def balance_signed(balance: Optional[float], bal_type: Optional[str]) -> Optional[float]:
    if balance is None:
        return None
    if bal_type and bal_type.strip().lower() == "dr":
        return -balance
    return balance


def infer_dr_cr_from_balance(txns: List[Txn], amount_tolerance: float = 0.50) -> List[Txn]:
    if len(txns) < 2:
        return txns

    balances_present = sum(1 for t in txns if t.balance is not None)
    if balances_present < 2:
        return txns

    # Work on a date-sorted view to handle multi-page tables that may be out of order.
    indexed = list(enumerate(txns))
    sorted_txns = sorted(indexed, key=lambda x: (x[1].date, x[0]))

    def best_match_amount(candidates: List[float], target: float) -> Optional[float]:
        if not candidates:
            return None
        best = min(candidates, key=lambda x: abs(x - target))
        rel_tol = max(amount_tolerance, target * 0.005)
        if abs(best - target) <= rel_tol:
            return best
        return None

    for (_prev_i, prev), (_cur_i, t) in zip(sorted_txns, sorted_txns[1:]):
        cur_bal = balance_signed(t.balance, t.balance_type)
        prev_bal = balance_signed(prev.balance, prev.balance_type)

        if cur_bal is None or prev_bal is None:
            continue

        delta = cur_bal - prev_bal
        if abs(delta) < 1e-9:
            continue

        amt = abs(delta)
        matched = best_match_amount(t.raw_amounts, amt)
        amt_final = matched if matched is not None else amt

        if delta > 0:
            # Credit movement: force credit, clear debit to avoid both columns being populated
            t.deposit_cr = amt_final
            t.withdrawal_dr = None
            t.inferred_type = "CR"
        else:
            # Debit movement
            t.withdrawal_dr = amt_final
            t.deposit_cr = None
            t.inferred_type = "DR"

    return txns


# -----------------------------
# Excel writing
# -----------------------------
def txns_to_dataframe(txns: List[Txn]) -> pd.DataFrame:
    cols = ["Date", "Narration", "Debit", "Credit", "Running Balance", "BalanceType", "InferredType", "Source", "RawAmounts", "RawText"]
    rows: List[Dict[str, Any]] = []
    for t in txns:
        rows.append(
            {
                "Date": t.date,
                "Narration": t.narration,
                "Debit": t.withdrawal_dr,
                "Credit": t.deposit_cr,
                "Running Balance": t.balance,
                "BalanceType": t.balance_type or "",
                "InferredType": t.inferred_type or "",
                "Source": t.source,
                "RawAmounts": ", ".join(f"{x:.2f}" for x in t.raw_amounts) if t.raw_amounts else "",
                "RawText": t.raw_text,
            }
        )
    # Ensure expected columns exist even when no rows are present to avoid KeyError later.
    return pd.DataFrame(rows, columns=cols)


def write_excel(out_path: str, meta: Dict[str, Optional[str]], diag: Dict[str, Any], df_txn: pd.DataFrame, raw_lines: Optional[List[str]] = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side

    wb = Workbook()

    ws = wb.active
    ws.title = "Transactions"
    cols = ["Date", "Narration", "Debit", "Credit", "Running Balance"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for r_idx, (_, r) in enumerate(df_txn[cols].iterrows(), start=2):
        for c_idx, name in enumerate(cols, start=1):
            val = r[name]
            if isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if name == "Date" and val is not None:
                cell.number_format = "dd/mm/yyyy"
            elif name in {"Debit", "Credit", "Running Balance"}:
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(vertical="top", wrap_text=(name == "Narration"))

    widths = {"A": 12, "B": 75, "C": 16, "D": 16, "E": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    if raw_lines:
        ws_raw = wb.create_sheet("Raw")
        ws_raw["A1"] = "Extracted lines (debug)"
        ws_raw["A1"].font = Font(bold=True)
        ws_raw.column_dimensions["A"].width = 120
        for i, ln in enumerate(raw_lines, start=2):
            ws_raw.cell(row=i, column=1, value=ln)

    wb.save(out_path)


# -----------------------------
# Orchestration
# -----------------------------
def parse_with_tables(pdf_path: str, which: str) -> Tuple[List[Txn], Dict[str, Any], List[str]]:
    if which == "camelot":
        tables = try_camelot_tables(pdf_path)
    elif which == "tabula":
        tables = try_tabula_tables(pdf_path)
    else:
        tables = []

    dfs = [t for t in tables if looks_like_txn_table(t)]
    txns_all: List[Txn] = []
    for df in dfs:
        txns_all.extend(normalize_table_to_txns(df, source=which))

    txns_all = infer_dr_cr_from_balance(txns_all)

    diag = {
        "method": which,
        "tables_found": len(tables),
        "tables_used": len(dfs),
        "txns_parsed": len(txns_all),
    }
    return txns_all, diag, []


def auto_parse(pdf_path: str, ocr: bool) -> Tuple[List[Txn], Dict[str, Any], List[str]]:
    candidates: List[Tuple[List[Txn], Dict[str, Any], List[str]]] = []

    camelot_txns, camelot_diag, camelot_raw = parse_with_tables(pdf_path, "camelot")
    if camelot_diag.get("txns_parsed", 0) >= 5:
        candidates.append((camelot_txns, camelot_diag, camelot_raw))

    tabula_txns, tabula_diag, tabula_raw = parse_with_tables(pdf_path, "tabula")
    if tabula_diag.get("txns_parsed", 0) >= 5:
        candidates.append((tabula_txns, tabula_diag, tabula_raw))

    lines_txns, lines_diag, lines_raw = parse_lines_method(pdf_path)
    if lines_diag.get("txns_parsed", 0) >= 5:
        candidates.append((lines_txns, lines_diag, lines_raw))

    if ocr:
        try:
            ocr_txns, ocr_diag, ocr_raw = parse_ocr_method(pdf_path)
            if ocr_diag.get("txns_parsed", 0) >= 5:
                candidates.append((ocr_txns, ocr_diag, ocr_raw))
        except Exception as e:
            candidates.append(([], {"method": "ocr", "error": str(e), "txns_parsed": 0}, []))

    if not candidates:
        return lines_txns, lines_diag, lines_raw

    def score(item: Tuple[List[Txn], Dict[str, Any], List[str]]) -> Tuple[int, int]:
        txns, diag, _ = item
        method = diag.get("method", "")
        pref = 2 if method in ("camelot", "tabula") else 1
        return (len(txns), pref)

    best = max(candidates, key=score)
    return best


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert a bank statement PDF to Excel (universal)")
    parser.add_argument("pdf", help="Input PDF path")
    parser.add_argument("xlsx", help="Output XLSX path")
    parser.add_argument(
        "--method",
        choices=["auto", "camelot", "tabula", "lines", "ocr"],
        default="auto",
        help="Extraction method to use (default: auto)",
    )
    parser.add_argument(
        "--ocr",
        action="store_true",
        help="Enable OCR fallback in auto mode (or use --method ocr)",
    )
    parser.add_argument(
        "--no-raw",
        action="store_true",
        help="Do not include the Raw debug sheet",
    )
    args = parser.parse_args()

    pdf_path = args.pdf
    out_xlsx = args.xlsx

    if not os.path.exists(pdf_path):
        print(f"ERROR: input PDF not found: {pdf_path}", file=sys.stderr)
        sys.exit(2)

    try:
        lines_for_meta = extract_lines_pdfplumber(pdf_path)
        full_text = "\n".join(lines_for_meta)
    except Exception:
        full_text = ""

    meta = extract_metadata(full_text)

    if args.method == "auto":
        txns, diag, raw_lines = auto_parse(pdf_path, ocr=args.ocr)
    elif args.method in ("camelot", "tabula"):
        txns, diag, raw_lines = parse_with_tables(pdf_path, args.method)
        if len(txns) < 3:
            txns, diag, raw_lines = parse_lines_method(pdf_path)
    elif args.method == "lines":
        txns, diag, raw_lines = parse_lines_method(pdf_path)
    elif args.method == "ocr":
        txns, diag, raw_lines = parse_ocr_method(pdf_path)
    else:
        txns, diag, raw_lines = parse_lines_method(pdf_path)

    df_txn = txns_to_dataframe(txns)

    if args.no_raw:
        raw_lines = []

    write_excel(out_xlsx, meta, diag, df_txn, raw_lines=raw_lines if raw_lines else None)
    print(f"Saved: {out_xlsx} (rows: {len(df_txn)})")


if __name__ == "__main__":
    main()
